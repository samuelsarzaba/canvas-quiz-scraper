from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.alert import Alert
import time

quiz_link = ""


def login(driver, wait, username, password):
    login_to_canvas_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//a[@title="Login to Canvas"]'))
    )
    login_to_canvas_button.click()

    username_field = wait.until(
        EC.presence_of_element_located((By.ID, "username"))
    )
    username_field.send_keys(username)

    password_field = driver.find_element(By.ID, "password")
    password_field.send_keys(password)

    login_button = driver.find_element(
        By.XPATH, '//button[@type="submit" and text()="Login"]'
    )
    login_button.click()

    button = wait.until(
        EC.element_to_be_clickable((By.ID, "dont-trust-browser-button"))
    )
    button.click()


def scrape_quiz(wait):
    time.sleep(3)

    question_element_container = wait.until(EC.presence_of_all_elements_located((By.ID, 'questions')))
    question_elements = question_element_container[0].find_elements(by=By.CLASS_NAME, value='display_question')

    question_texts = [question_element.find_element(by=By.CLASS_NAME, value='text') for question_element in question_elements]

    question_names = [question_text.find_element(by=By.CLASS_NAME, value='question_text') for question_text in question_texts]
    question_answers = [question_text.find_element(by=By.CLASS_NAME, value='answers') for question_text in question_texts]
    correct_question_answers = [question_answer.find_elements(by=By.CLASS_NAME, value='correct_answer') for question_answer in question_answers]

    questions_dict = {}

    for question_name, correct_answer in zip(question_names, correct_question_answers):
        correct_answer_texts = []
        for answer in correct_answer:
            correct_answer_texts.append(answer.text.strip())
        questions_dict[question_name.text] = correct_answer_texts

    return questions_dict


def save_dict_to_excel(data_dict, filename='output.xlsx'):
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write headers in the first column
    headers = list(data_dict.keys())
    for row, header in enumerate(headers, start=1):
        ws.cell(row=row, column=1, value=header)

    # Write data
    for row, (key, values) in enumerate(data_dict.items(), start=1):
        for col, value in enumerate(values, start=2):
            ws.cell(row=row, column=col, value=value)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(filename)


def main():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    wait = WebDriverWait(driver, 10)

    try:
        driver.get(quiz_link)

        login(driver, wait, "", "")

        save_dict_to_excel(scrape_quiz(wait))

    finally:
        driver.quit()


if __name__ == "__main__":
    main()
